"""Generate the SAPHFIRE injection-prep fill-in form.

Writes SAPHFIRE_injection_prep_form.xlsx with four sheets:
  Read me        - instructions + legend
  Gas bottles    - one row per gas-phase species; you fill pressure / purity /
                   method / syringe etc.; reference targets are pre-filled
  Liquid mixtures- Solutions A-D; you fill syringe availability / vial / density
  Scenario       - the assumptions behind the reference numbers

Yellow cells = you fill in. White cells = reference (pre-computed). Once filled,
hand it back and the planner finalizes all gas-phase volumes / # injections /
minutes and the campaign sufficiency check.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from injection_model import SOLUTION_NAMES, load_planner

OUT = Path(__file__).with_name("SAPHFIRE_injection_prep_form.xlsx")
TOTAL_VOC = 150.0
N_EXP = 15
SAFETY = 1.5

FILL = PatternFill("solid", fgColor="FFF2CC")        # yellow: user fills
HEAD = PatternFill("solid", fgColor="4472C4")        # blue header
REFH = PatternFill("solid", fgColor="D9E1F2")        # light-blue reference header
WHITE_BOLD = Font(bold=True, color="FFFFFF")
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _scenario_values(p):
    res = p.inverse_voc(TOTAL_VOC)
    gas_ppb = {r["key"]: r["ppb"] for _, r in res["gases"].iterrows()}
    co = p.co_from_voc(TOTAL_VOC)["co_ppb"]
    noy = {rg: TOTAL_VOC * p.noy_voc_ratio(p.noy_params["mce_presets"][rg])
           for rg in ("flaming", "smoldering")}
    spec = p.noy_params["noy_speciation"]
    per_exp = {}  # key -> (flaming ppb, smoldering ppb)
    for k, v in gas_ppb.items():
        per_exp[k] = (v, v)
    per_exp["co"] = (co, co)
    for sp, frac in spec.items():
        per_exp[sp] = (noy["flaming"] * frac, noy["smoldering"] * frac)
    sol_uL = {r["solution"]: r["inject_volume_uL"] for _, r in res["per_solution"].iterrows()}
    return per_exp, sol_uL


def _style_header(ws, row, headers, fill=HEAD, font=WHITE_BOLD):
    for c, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=text)
        cell.fill = fill
        cell.font = font
        cell.alignment = CENTER
        cell.border = BORDER


def build():
    p = load_planner()
    per_exp, sol_uL = _scenario_values(p)
    mw = dict(zip(p.bottles["key"], p.bottles["MW_g_per_mol"]))
    voc_mw = dict(zip(p.voc["key"], p.voc["MW_g_per_mol"]))
    mw = {**voc_mw, **mw}

    wb = Workbook()

    # --- Read me ---
    ws = wb.active
    ws.title = "Read me"
    ws["A1"] = "SAPHFIRE 2026 — injection preparation form"
    ws["A1"].font = Font(bold=True, size=14)
    notes = [
        "",
        "Fill in the YELLOW cells. White cells are reference values already computed.",
        "",
        "Gas bottles sheet — for each species record:",
        "   • bottle ID/size/current pressure (bar)",
        "   • status: pure or mixture; if mixture, the concentration (value + unit % or ppm) and balance gas",
        "   • injection method: MFC or volume injection",
        "       - if MFC: the flow used (available MFC max flows: 0.5 / 5 / 20 LPM)",
        "       - if volume injection: the loop volume (0.32 or 3.2 mL)",
        "   • whether a regulator is available",
        "   Note: volume injection works for pure/high-% bottles (0.32 mL≈1.2 ppb, "
        "3.2 mL≈12 ppb if pure); use an MFC for dilute (ppm) cylinders.",
        "",
        "Liquid mixtures sheet — for each solution (A-D) record:",
        "   • whether it's prepared, the vial volume and measured density",
        "   • the syringe available for injection and whether it suits the per-injection µL shown",
        "",
        f"Reference scenario: total VOC = {TOTAL_VOC:.0f} ppb, {N_EXP} experiments, "
        f"safety factor {SAFETY:g} (see Scenario sheet).",
        "",
        "Once filled, send it back and the planner finalizes: gas-phase volumes / # injections /",
        "minutes per species, whether each bottle lasts the campaign, and the final export plan.",
    ]
    for i, t in enumerate(notes, start=2):
        ws.cell(row=i, column=1, value=t)
    ws.cell(row=3, column=1).fill = FILL
    ws.column_dimensions["A"].width = 95

    # --- Gas bottles ---
    gw = wb.create_sheet("Gas bottles")
    ref_cols = ["species", "MW\n(g/mol)", "per-exp ppb\n(flaming)",
                "per-exp ppb\n(smoldering)", f"campaign ppb\n(×{N_EXP}×{SAFETY:g}, flaming)"]
    fill_cols = ["bottle ID /\nlabel", "bottle size\n(L)", "current\npressure (bar)",
                 "status\n(pure/mixture)", "concentration\n(value)", "conc. unit\n(%/ppm)",
                 "balance\ngas", "injection method\n(MFC / volume)",
                 "volume inj. loop\n(0.32 / 3.2 mL)", "MFC flow\n(0.5/5/20 LPM)",
                 "regulator\n(Y/N)", "notes"]
    headers = ref_cols + fill_cols
    _style_header(gw, 1, ref_cols, fill=REFH, font=BOLD)
    for c in range(len(ref_cols) + 1, len(headers) + 1):
        cell = gw.cell(row=1, column=c, value=headers[c - 1])
        cell.fill = HEAD; cell.font = WHITE_BOLD; cell.alignment = CENTER; cell.border = BORDER

    species = ["propene", "ethene", "acetylene", "co", "no", "no2", "hono"]
    labels = {"propene": "Propene", "ethene": "Ethene", "acetylene": "Acetylene",
              "co": "CO", "no": "NO", "no2": "NO2", "hono": "HONO"}
    for r, k in enumerate(species, start=2):
        fl, sm = per_exp.get(k, (0.0, 0.0))
        gw.cell(row=r, column=1, value=labels[k])
        gw.cell(row=r, column=2, value=round(mw.get(k, float("nan")), 3))
        gw.cell(row=r, column=3, value=round(fl, 2))
        gw.cell(row=r, column=4, value=round(sm, 2))
        gw.cell(row=r, column=5, value=round(fl * N_EXP * SAFETY, 1))
        for c in range(1, len(headers) + 1):
            cell = gw.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = CENTER if c <= len(ref_cols) else WRAP
            if c > len(ref_cols):
                cell.fill = FILL

    # dropdowns
    n = len(species) + 1
    _dv(gw, '"pure,mixture"', f"I2:I{n}")
    _dv(gw, '"%,ppm"', f"K2:K{n}")
    _dv(gw, '"MFC,volume injection"', f"M2:M{n}")
    _dv(gw, '"n/a,0.32,3.2"', f"N2:N{n}")     # n/a = not this method
    _dv(gw, '"n/a,0.5,5,20"', f"O2:O{n}")
    _dv(gw, '"Y,N"', f"P2:P{n}")
    _widths(gw, [11, 8, 11, 12, 13, 11, 9, 11, 13, 12, 10, 9, 13, 14, 14, 9, 22])
    gw.freeze_panes = "A2"

    # --- Liquid mixtures ---
    lw = wb.create_sheet("Liquid mixtures")
    ref = ["solution", "mixture", "per-exp inject\n(µL)", f"campaign total\n(mL, ×{N_EXP}×{SAFETY:g})"]
    fillc = ["prepared?\n(Y/N)", "vial volume\n(mL)", "measured density\n(g/mL)",
             "syringe available\n(type / volume)", "suits per-exp µL?\n(Y/N)",
             "injection port /\nmethod", "notes"]
    headers = ref + fillc
    _style_header(lw, 1, ref, fill=REFH, font=BOLD)
    for c in range(len(ref) + 1, len(headers) + 1):
        cell = lw.cell(row=1, column=c, value=headers[c - 1])
        cell.fill = HEAD; cell.font = WHITE_BOLD; cell.alignment = CENTER; cell.border = BORDER
    for r, s in enumerate("ABCD", start=2):
        uL = sol_uL.get(s, float("nan"))
        lw.cell(row=r, column=1, value=s)
        lw.cell(row=r, column=2, value=SOLUTION_NAMES[s])
        lw.cell(row=r, column=3, value=round(uL, 2))
        lw.cell(row=r, column=4, value=round(uL * N_EXP * SAFETY / 1000.0, 3))
        for c in range(1, len(headers) + 1):
            cell = lw.cell(row=r, column=c)
            cell.border = BORDER
            cell.alignment = CENTER if c <= len(ref) else WRAP
            if c > len(ref):
                cell.fill = FILL
    _dv(lw, '"Y,N"', "E2:E5")
    _dv(lw, '"Y,N"', "I2:I5")
    _widths(lw, [9, 22, 13, 15, 10, 11, 15, 18, 13, 16, 22])
    lw.freeze_panes = "A2"

    # --- Scenario ---
    sw = wb.create_sheet("Scenario")
    sw["A1"] = "Reference scenario / assumptions"; sw["A1"].font = Font(bold=True, size=12)
    rows = [
        ("Chamber", p.chamber.name),
        ("Volume (m³)", p.chamber.volume_m3),
        ("Temperature (K)", p.chamber.temperature_K),
        ("Pressure (Pa)", p.chamber.pressure_Pa),
        ("n_air (mol)", round(p.chamber.n_air)),
        ("Total VOC target (ppb)", TOTAL_VOC),
        ("Number of experiments", N_EXP),
        ("Safety factor", SAFETY),
        ("MCE flaming", p.noy_params["mce_presets"]["flaming"]),
        ("MCE smoldering", p.noy_params["mce_presets"]["smoldering"]),
        ("NOy:VOC source", "Gkatzelis 2024 ACP Fig. 7c (power fit)"),
        ("NOy speciation", "70% NO / 20% NO2 / 10% HONO (Roberts 2020)"),
        ("CO:VOC source", "Gkatzelis 2024 ACP Fig. 7b"),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        sw.cell(row=i, column=1, value=k).font = BOLD
        sw.cell(row=i, column=2, value=v)
    sw.column_dimensions["A"].width = 26
    sw.column_dimensions["B"].width = 44

    wb.save(OUT)
    return OUT


def _dv(ws, formula, cell_range):
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(cell_range)


def _widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


if __name__ == "__main__":
    path = build()
    print(f"Wrote {path}")
